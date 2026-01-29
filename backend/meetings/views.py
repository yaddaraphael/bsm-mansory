# backend/meetings/views.py
from __future__ import annotations

import io
import logging
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db import transaction
from django.db.models import Q, Count, Prefetch, Exists, OuterRef
from django.shortcuts import get_object_or_404
from django.http import HttpResponse

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Meeting, MeetingJob, MeetingJobPhase
from .permissions import MeetingPermission
from .serializers import (
    MeetingSerializer,
    MeetingListSerializer,
    MeetingJobSerializer,
    MeetingJobCreateUpdateSerializer,
    MeetingJobPhaseSerializer,
    MeetingJobPhaseCreateUpdateSerializer,
)

from projects.models import Project
from branches.models import Branch

logger = logging.getLogger(__name__)


def _normalize_scope_key(value: str) -> str:
    if not value:
        return ""
    return "".join(ch for ch in value.upper() if ch.isalnum())


def _to_decimal(value) -> Decimal:
    try:
        return Decimal(str(value if value is not None else 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


# ---------------------------------------------------------
# Helper: role filtering
# ---------------------------------------------------------
def filter_meetings_by_role(queryset, user):
    """
    Apply role-based visibility rules.
    """
    if user.role in ["ROOT_SUPERADMIN", "SUPERADMIN", "ADMIN"]:
        return queryset

    if user.role == "PROJECT_MANAGER":
        return queryset.filter(meeting_jobs__project__project_manager=user).distinct()

    if user.role == "BRANCH_MANAGER":
        if user.division:
            return queryset.filter(
                Q(branch=user.division) | Q(meeting_jobs__project__branch=user.division)
            ).distinct()
        return queryset.none()

    return queryset.none()


# ---------------------------------------------------------
# ViewSet
# ---------------------------------------------------------
class MeetingViewSet(viewsets.ModelViewSet):
    """
    Meetings API.

    Performance strategy:
    - list(): lightweight serializer + minimal prefetch
    - retrieve(): full serializer + prefetch jobs + phases + project relations
    """

    permission_classes = [IsAuthenticated, MeetingPermission]
    ordering = ["-meeting_date", "-created_at"]

    # ---- Serializer selection (FAST list, full detail) ----
    def get_serializer_class(self):
        if self.action == "list":
            return MeetingListSerializer
        return MeetingSerializer

    # ---- Querysets per action ----
    def get_queryset(self):
        user = self.request.user

        # Base queryset (light)
        qs = Meeting.objects.select_related("created_by", "branch").annotate(
            meeting_jobs_count=Count("meeting_jobs", distinct=True)
        )

        # Role filter
        qs = filter_meetings_by_role(qs, user)

        # Query param filters (list view only)
        if self.action == "list":
            params = self.request.query_params
            status_param = (params.get("status") or "").upper()
            if status_param in ["DRAFT", "COMPLETED"]:
                qs = qs.filter(status=status_param)

            branch_param = params.get("branch")
            if branch_param:
                qs = qs.filter(branch_id=branch_param)

            date_from = params.get("date_from")
            if date_from:
                qs = qs.filter(meeting_date__gte=date_from)

            date_to = params.get("date_to")
            if date_to:
                qs = qs.filter(meeting_date__lte=date_to)

            search = (params.get("search") or "").strip()
            if search:
                qs = qs.filter(
                    Q(notes__icontains=search)
                    | Q(created_by__username__icontains=search)
                    | Q(created_by__first_name__icontains=search)
                    | Q(created_by__last_name__icontains=search)
                    | Q(branch__name__icontains=search)
                )

        # If detail/retrieve or export or jobs endpoints need deeper data:
        detail_actions = {
            "retrieve",
            "jobs",
            "job_phases",
            "export_pdf",
            "export_excel",
            "batch_save_jobs",
        }
        if self.action in detail_actions:
            qs = qs.prefetch_related(
                Prefetch(
                    "meeting_jobs",
                    queryset=MeetingJob.objects.select_related(
                        "project",
                        "project__project_manager",
                        "project__foreman",
                        "project__branch",
                    )
                    .prefetch_related("phases", "project__scopes")
                    .all(),
                )
            )

        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    # ---------------------------------------------------------
    # /meetings/{id}/jobs/
    # ---------------------------------------------------------
    @action(detail=True, methods=["get", "post", "put", "patch"])
    def jobs(self, request, pk=None):
        meeting = self.get_object()

        if request.method == "GET":
            qs = meeting.meeting_jobs.select_related("project").prefetch_related("phases")
            return Response(MeetingJobSerializer(qs, many=True).data)

        if request.method == "POST":
            data = request.data.copy()
            data["meeting"] = meeting.id
            serializer = MeetingJobCreateUpdateSerializer(
                data=data,
                context={"request": request},
            )
            serializer.is_valid(raise_exception=True)
            serializer.save(meeting=meeting)
            # Return in read format (with project + phases)
            mj = MeetingJob.objects.select_related("project").prefetch_related("phases").get(
                id=serializer.instance.id
            )
            return Response(MeetingJobSerializer(mj).data, status=status.HTTP_201_CREATED)

        # PUT/PATCH
        job_id = request.data.get("id")
        if not job_id:
            return Response({"detail": "Job entry ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        meeting_job = get_object_or_404(MeetingJob, id=job_id, meeting=meeting)
        serializer = MeetingJobCreateUpdateSerializer(
            meeting_job,
            data=request.data,
            partial=(request.method == "PATCH"),
            context={"request": request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()

        mj = MeetingJob.objects.select_related("project").prefetch_related("phases").get(id=meeting_job.id)
        return Response(MeetingJobSerializer(mj).data, status=status.HTTP_200_OK)

    # ---------------------------------------------------------
    # /meetings/{id}/jobs/{job_id}/phases/
    # ---------------------------------------------------------
    @action(
        detail=True,
        methods=["get", "post", "put", "patch", "delete"],
        url_path=r"jobs/(?P<job_id>[^/.]+)/phases",
    )
    def job_phases(self, request, pk=None, job_id=None):
        meeting = self.get_object()
        meeting_job = get_object_or_404(MeetingJob, id=job_id, meeting=meeting)

        if request.method == "GET":
            phases = meeting_job.phases.all().order_by("phase_code")
            return Response(MeetingJobPhaseSerializer(phases, many=True).data)

        if request.method == "POST":
            data = request.data.copy()
            data["meeting_job"] = meeting_job.id
            serializer = MeetingJobPhaseCreateUpdateSerializer(data=data)
            serializer.is_valid(raise_exception=True)
            serializer.save(meeting_job=meeting_job)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        if request.method in ["PUT", "PATCH"]:
            phase_id = request.data.get("id")
            if not phase_id:
                return Response({"detail": "Phase ID is required"}, status=status.HTTP_400_BAD_REQUEST)

            phase = get_object_or_404(MeetingJobPhase, id=phase_id, meeting_job=meeting_job)
            serializer = MeetingJobPhaseCreateUpdateSerializer(
                phase, data=request.data, partial=(request.method == "PATCH")
            )
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)

        # DELETE
        phase_id = request.query_params.get("id")
        if not phase_id:
            return Response({"detail": "Phase ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        phase = get_object_or_404(MeetingJobPhase, id=phase_id, meeting_job=meeting_job)
        phase.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # ---------------------------------------------------------
    # /meetings/meetings/active_jobs/
    # ---------------------------------------------------------
    @action(detail=False, methods=["get"])
    def active_jobs(self, request):
        """
        Fast active jobs:
        - only ACTIVE projects
        - only projects that exist in SpectrumJob table (via EXISTS subquery)
        - role filters match your earlier behavior
        """
        user = request.user

        from spectrum.models import SpectrumJob  # local import to avoid cycles

        spectrum_exists = SpectrumJob.objects.filter(job_number=OuterRef("job_number"))

        projects = (
            Project.objects.filter(status="ACTIVE")
            .exclude(job_number__isnull=True)
            .exclude(job_number="")
            .annotate(in_spectrum=Exists(spectrum_exists))
            .filter(in_spectrum=True)
            .select_related("branch", "project_manager", "foreman")
            .distinct()
        )

        if user.role == "PROJECT_MANAGER":
            projects = projects.filter(project_manager=user)
        elif user.role == "BRANCH_MANAGER":
            if user.division:
                projects = projects.filter(branch=user.division)
            else:
                user_branch = Branch.objects.filter(manager=user).first()
                projects = projects.filter(branch=user_branch) if user_branch else projects.none()

        include_scopes = request.query_params.get("include_scopes") == "1"
        if include_scopes:
            projects = projects.prefetch_related("scopes")
            from projects.serializers import ProjectSerializer
            return Response(ProjectSerializer(projects, many=True).data)

        from projects.serializers import ProjectListSerializer
        return Response(ProjectListSerializer(projects, many=True).data)

    # ---------------------------------------------------------
    # /meetings/meetings/project_phases/?project_id=123
    # ---------------------------------------------------------
    @action(detail=False, methods=["get"], url_path="project_phases")
    def project_phases(self, request):
        project_id = request.query_params.get("project_id")
        if not project_id:
            return Response({"detail": "project_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        phases = (
            MeetingJobPhase.objects.filter(
                meeting_job__project_id=project_id,
                meeting_job__meeting__status="COMPLETED",
            )
            .select_related("meeting_job__meeting")
            .order_by("meeting_job__meeting__meeting_date", "updated_at")
        )

        data = [
            {
                "phase_code": ph.phase_code,
                "phase_description": ph.phase_description,
                "quantity": ph.quantity,
                "installed_quantity": ph.installed_quantity,
                "percent_complete": ph.percent_complete,
                "meeting_date": ph.meeting_job.meeting.meeting_date,
                "updated_at": ph.updated_at,
            }
            for ph in phases
        ]

        return Response({"phases": data})

    # ---------------------------------------------------------
    # /meetings/meetings/batch_job_details/
    # ---------------------------------------------------------
    @action(detail=False, methods=["post"])
    def batch_job_details(self, request):
        """
        Batch job details (dates + phases + scopes).
        Optional meeting_id enables previous meeting calculations.
        """
        job_numbers = request.data.get("job_numbers", [])
        if not isinstance(job_numbers, list):
            return Response({"detail": "job_numbers array is required"}, status=status.HTTP_400_BAD_REQUEST)

        job_numbers = [jn for jn in job_numbers if jn]
        if not job_numbers:
            return Response({})

        meeting_id = request.data.get("meeting_id")

        from spectrum.models import SpectrumJobDates, SpectrumPhaseEnhanced
        from projects.serializers import ProjectScopeSerializer

        # ---- meeting maps (previous + current) ----
        prev_installed_map: dict[tuple[int, str], Decimal] = {}
        prev_meeting_installed_map: dict[tuple[int, str], Decimal] = {}

        def build_installed_map(*, meeting: Meeting | None = None, before_meeting: Meeting | None = None) -> dict[tuple[int, str], Decimal]:
            latest_map: dict[tuple[int, str], tuple[date, datetime, Decimal]] = {}
            phases_qs = (
                MeetingJobPhase.objects.filter(
                    meeting_job__project__job_number__in=job_numbers,
                    meeting_job__meeting__status="COMPLETED",
                )
                .select_related("meeting_job__project", "meeting_job__meeting")
            )
            if meeting is not None:
                phases_qs = phases_qs.filter(meeting_job__meeting=meeting)
            if before_meeting is not None:
                phases_qs = phases_qs.filter(
                    Q(meeting_job__meeting__meeting_date__lt=before_meeting.meeting_date)
                    | Q(
                        meeting_job__meeting__meeting_date=before_meeting.meeting_date,
                        meeting_job__meeting__id__lt=before_meeting.id,
                    )
                )

            for ph in phases_qs:
                key = (ph.meeting_job.project_id, _normalize_scope_key(ph.phase_code or ""))
                if not key[1]:
                    continue
                meeting_date = ph.meeting_job.meeting.meeting_date or date.min
                updated_at = ph.updated_at or datetime.min
                qty = _to_decimal(ph.installed_quantity)

                current = latest_map.get(key)
                if not current or (meeting_date, updated_at) > (current[0], current[1]):
                    latest_map[key] = (meeting_date, updated_at, qty)

            return {k: v[2] for k, v in latest_map.items()}

        if meeting_id:
            current_meeting = get_object_or_404(Meeting, id=meeting_id)
            previous_meeting = (
                Meeting.objects.filter(
                    meeting_date__lt=current_meeting.meeting_date,
                    status="COMPLETED",
                )
                .order_by("-meeting_date")
                .first()
            )
            prev_installed_map = build_installed_map(before_meeting=current_meeting)
            prev_meeting_installed_map = build_installed_map(meeting=previous_meeting)

        # ---- Spectrum dates ----
        dates_qs = SpectrumJobDates.objects.filter(job_number__in=job_numbers)
        dates_by_job = {d.job_number: d for d in dates_qs}

        # ---- Spectrum phases ----
        phases_qs = SpectrumPhaseEnhanced.objects.filter(job_number__in=job_numbers)
        phases_by_job: dict[str, list[dict]] = {}
        for ph in phases_qs:
            phases_by_job.setdefault(ph.job_number, []).append(
                {
                    "phase_code": ph.phase_code,
                    "description": ph.description,
                    "jtd_quantity": ph.jtd_quantity,
                    "estimated_quantity": ph.estimated_quantity,
                    "start_date": ph.start_date,
                    "end_date": ph.end_date,
                }
            )

        # ---- Project scopes (with optional previous values) ----
        projects = (
            Project.objects.filter(job_number__in=job_numbers)
            .prefetch_related("scopes__scope_type")
        )
        projects_by_job = {p.job_number: p for p in projects}

        def scope_maps_for(project_id: int, scope_code: str, scope_name: str):
            norm_code = _normalize_scope_key(scope_code)
            norm_name = _normalize_scope_key(scope_name)
            return norm_code, norm_name

        def lookup_installed(
            installed_map: dict[tuple[int, str], Decimal],
            project_id: int,
            norm_code: str,
            norm_name: str,
        ) -> Decimal:
            val = installed_map.get((project_id, norm_code))
            if val is not None:
                return val
            val = installed_map.get((project_id, norm_name))
            if val is not None:
                return val
            if norm_code or norm_name:
                for (pid, key), qty in installed_map.items():
                    if pid != project_id:
                        continue
                    if norm_code and norm_code in key:
                        return qty
                    if norm_name and norm_name in key:
                        return qty
            return Decimal("0")

        # ---- Build response ----
        payload: dict[str, dict] = {}
        for job_number in job_numbers:
            details: dict[str, object] = {}

            dates = dates_by_job.get(job_number)
            if dates:
                details["dates"] = {
                    "start_date": dates.start_date,
                    "est_start_date": dates.est_start_date,
                    "complete_date": dates.complete_date,
                    "projected_complete_date": dates.projected_complete_date,
                    "est_complete_date": dates.est_complete_date,
                }

            if job_number in phases_by_job:
                details["phases"] = phases_by_job[job_number]

            project = projects_by_job.get(job_number)
            if project:
                scopes = ProjectScopeSerializer(project.scopes.all(), many=True).data

                if meeting_id:
                    for scope in scopes:
                        scope_type_detail = scope.get("scope_type_detail") or {}
                        scope_type_obj = scope.get("scope_type") if isinstance(scope.get("scope_type"), dict) else {}
                        scope_code = scope_type_obj.get("code") or scope_type_detail.get("code") or ""
                        scope_name = scope_type_obj.get("name") or scope_type_detail.get("name") or ""
                        norm_code, norm_name = scope_maps_for(project.id, scope_code, scope_name)

                        prev_installed = lookup_installed(prev_meeting_installed_map, project.id, norm_code, norm_name)
                        installed_before = lookup_installed(prev_installed_map, project.id, norm_code, norm_name)

                        scope_qty = _to_decimal(scope.get("quantity", scope.get("qty_sq_ft", 0)))
                        previous_balance = max(Decimal("0"), scope_qty - installed_before)

                        # Use cumulative installed before this meeting as the baseline for UI math
                        scope["previous_meeting_installed"] = installed_before
                        scope["previous_balance"] = previous_balance
                        scope["last_meeting_installed"] = prev_installed

                details["scopes"] = scopes

            payload[job_number] = details

        return Response(payload)

    # ---------------------------------------------------------
    # /meetings/{id}/batch_save_jobs/
    # ---------------------------------------------------------
    @action(detail=True, methods=["post"])
    def batch_save_jobs(self, request, pk=None):
        """
        Batch save jobs + phases (optimized, safe, avoids duplicates).
        """
        meeting = self.get_object()
        jobs_data = request.data.get("jobs", [])

        if not jobs_data or not isinstance(jobs_data, list):
            return Response({"detail": "jobs array is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                saved_job_ids = []

                # Preload existing meeting jobs by project_id for fewer queries
                existing_jobs = {
                    mj.project_id: mj
                    for mj in MeetingJob.objects.filter(meeting=meeting).select_related("project")
                }

                for job_data in jobs_data:
                    project_id = job_data.get("project_id")
                    if not project_id:
                        continue

                    mj = existing_jobs.get(project_id)

                    defaults = {
                        "masons": job_data.get("masons", 0),
                        "labors": job_data.get("labors", 0),
                        "notes": job_data.get("notes", "") or "",
                        "handoff_from_estimator": bool(job_data.get("handoff_from_estimator", False)),
                        "handoff_to_foreman": bool(job_data.get("handoff_to_foreman", False)),
                        "site_specific_safety_plan": bool(job_data.get("site_specific_safety_plan", False)),
                        "saturdays": job_data.get("saturdays"),
                        "full_weekends": job_data.get("full_weekends"),
                        "selected_scope": job_data.get("selected_scope", "") or "",
                    }

                    if mj is None:
                        mj, _ = MeetingJob.objects.update_or_create(
                            meeting=meeting,
                            project_id=project_id,
                            defaults=defaults,
                        )
                        existing_jobs[project_id] = mj
                    else:
                        for k, v in defaults.items():
                            setattr(mj, k, v)
                        mj.save(update_fields=list(defaults.keys()) + ["updated_at"])

                    # phases
                    phases_data = job_data.get("phases", []) or []
                    if phases_data:
                        existing_phases = {p.phase_code: p for p in mj.phases.all()}
                        processed = set()

                        for ph in phases_data:
                            code = (ph.get("phase_code") or "").strip()
                            if not code or code in processed:
                                continue
                            processed.add(code)

                            MeetingJobPhase.objects.update_or_create(
                                meeting_job=mj,
                                phase_code=code,
                                defaults={
                                    "phase_description": ph.get("phase_description", "") or "",
                                    "masons": ph.get("masons", 0) or 0,
                                    "operators": ph.get("operators", 0) or 0,
                                    "labors": ph.get("labors", 0) or 0,
                                    "quantity": ph.get("quantity", 0) or 0,
                                    "installed_quantity": ph.get("installed_quantity", 0) or 0,
                                    "duration": ph.get("duration"),
                                    "notes": ph.get("notes", "") or "",
                                },
                            )

                        # delete removed phases
                        to_delete = set(existing_phases.keys()) - processed
                        if to_delete:
                            MeetingJobPhase.objects.filter(meeting_job=mj, phase_code__in=to_delete).delete()

                    saved_job_ids.append(mj.id)

                # status / notifications
                is_draft = bool(request.data.get("is_draft", False))
                if not is_draft:
                    meeting.status = "COMPLETED"
                    meeting.save(update_fields=["status", "updated_at"])
                    self._send_meeting_notifications(meeting)
                    try:
                        from .signals import sync_meeting_phase_to_project_scope
                        phases = MeetingJobPhase.objects.filter(meeting_job__meeting=meeting)
                        for ph in phases:
                            sync_meeting_phase_to_project_scope(ph)
                    except Exception as e:
                        logger.error("Error syncing phases on meeting complete: %s", e, exc_info=True)

                saved = (
                    MeetingJob.objects.filter(id__in=saved_job_ids)
                    .select_related("project", "project__branch", "project__project_manager", "project__foreman")
                    .prefetch_related("phases", "project__scopes")
                )
                return Response(MeetingJobSerializer(saved, many=True).data, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error("Error batch saving jobs: %s", e, exc_info=True)
            return Response(
                {"detail": f"Error batch saving jobs: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # ---------------------------------------------------------
    # Notifications (kept, safe)
    # ---------------------------------------------------------
    def _send_meeting_notifications(self, meeting: Meeting):
        try:
            from accounts.models import Notification
            from django.contrib.contenttypes.models import ContentType

            meeting_ct = ContentType.objects.get_for_model(Meeting)
            meeting_jobs = meeting.meeting_jobs.select_related(
                "project", "project__project_manager", "project__branch"
            ).all()

            notified_pms = set()
            notified_bms = set()

            for job in meeting_jobs:
                project = job.project

                # PM
                if project.project_manager and project.project_manager_id not in notified_pms:
                    Notification.objects.create(
                        user=project.project_manager,
                        type="REPORT_SUBMITTED",
                        title=f"Meeting Report Available - {meeting.meeting_date}",
                        message=(
                            f"A meeting report for {project.job_number} - {project.name} is now available. "
                            f"The meeting was held on {meeting.meeting_date}."
                        ),
                        link=f"/meetings/{meeting.id}/review",
                        content_type=meeting_ct,
                        object_id=meeting.id,
                    )
                    notified_pms.add(project.project_manager_id)

                # Branch manager
                branch_manager = getattr(project.branch, "manager", None) if project.branch else None
                bm_id = getattr(project.branch, "manager_id", None) if project.branch else None
                if branch_manager and bm_id:
                    if bm_id not in notified_bms:
                        branch_jobs_count = meeting_jobs.filter(project__branch=project.branch).count()
                        Notification.objects.create(
                            user=branch_manager,
                            type="REPORT_SUBMITTED",
                            title=f"Meeting Report Available - {meeting.meeting_date}",
                            message=(
                                f"A meeting report for {project.branch.name} is now available with "
                                f"{branch_jobs_count} project(s). The meeting was held on {meeting.meeting_date}."
                            ),
                            link=f"/meetings/{meeting.id}/review",
                            content_type=meeting_ct,
                            object_id=meeting.id,
                        )
                        notified_bms.add(bm_id)

        except Exception as e:
            logger.error("Error sending meeting notifications: %s", e, exc_info=True)
            # do not fail

    # ---------------------------------------------------------
    # Export PDF
    # ---------------------------------------------------------
    @action(detail=True, methods=["get"])
    def export_pdf(self, request, pk=None):
        """
        Kept close to your original behavior, but uses the optimized queryset.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            return Response(
                {"detail": "PDF export requires reportlab library"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        meeting = self.get_object()
        user = request.user

        meeting_jobs = meeting.meeting_jobs.select_related(
            "project", "project__project_manager", "project__foreman", "project__branch"
        ).prefetch_related("phases", "project__scopes")

        # role filter for export
        if user.role == "BRANCH_MANAGER":
            if user.division:
                meeting_jobs = meeting_jobs.filter(project__branch=user.division)
            else:
                meeting_jobs = meeting_jobs.none()
        elif user.role == "PROJECT_MANAGER":
            meeting_jobs = meeting_jobs.filter(project__project_manager=user)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=30,
        )
        story.append(Paragraph(f"Meeting Report - {meeting.meeting_date}", title_style))
        story.append(Spacer(1, 0.2 * inch))

        info_data = [
            ["Meeting Date:", str(meeting.meeting_date)],
            ["Created By:", meeting.created_by.get_full_name() or meeting.created_by.username],
            ["Branch:", meeting.branch.name if meeting.branch else "All Branches"],
            ["Created At:", meeting.created_at.strftime("%Y-%m-%d %H:%M:%S")],
        ]
        info_table = Table(info_data, colWidths=[2 * inch, 4 * inch])
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        if meeting.notes:
            story.append(Paragraph("<b>Meeting Notes:</b>", styles["Heading2"]))
            story.append(Spacer(1, 0.1 * inch))
            story.append(Paragraph(meeting.notes.replace("\n", "<br/>"), styles["Normal"]))
            story.append(Spacer(1, 0.3 * inch))

        # Jobs table
        story.append(Paragraph("<b>Job Details:</b>", styles["Heading2"]))
        story.append(Spacer(1, 0.1 * inch))

        jobs_data = [[
            "Job Number", "Project Name", "Branch", "PM", "Foreman",
            "Sat", "Weekends", "Handoff Est", "Handoff Foreman", "Safety", "Masons", "Labors"
        ]]

        for mj in meeting_jobs:
            p = mj.project
            pm_name = p.project_manager.get_full_name() if p.project_manager else "N/A"
            foreman_name = p.foreman.get_full_name() if p.foreman else "N/A"
            jobs_data.append([
                p.job_number or "",
                (p.name or "")[:35],
                p.branch.name if p.branch else "N/A",
                pm_name,
                foreman_name,
                "Yes" if mj.saturdays is True else ("No" if mj.saturdays is False else "N/A"),
                "Yes" if mj.full_weekends is True else ("No" if mj.full_weekends is False else "N/A"),
                "Yes" if mj.handoff_from_estimator else "No",
                "Yes" if mj.handoff_to_foreman else "No",
                "Yes" if mj.site_specific_safety_plan else "No",
                str(mj.masons),
                str(mj.labors),
            ])

        jobs_table = Table(jobs_data)
        jobs_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 1, colors.grey),
                ]
            )
        )
        story.append(jobs_table)

        doc.build(story)
        buffer.seek(0)

        filename = f"meeting_{meeting.meeting_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        resp = HttpResponse(buffer.read(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ---------------------------------------------------------
    # Export Excel (kept simpler but stable)
    # ---------------------------------------------------------
    @action(detail=True, methods=["get"])
    def export_excel(self, request, pk=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as e:
            return Response(
                {"detail": f"Excel export requires openpyxl library. Error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        meeting = self.get_object()
        user = request.user

        meeting_jobs = meeting.meeting_jobs.select_related(
            "project", "project__project_manager", "project__foreman", "project__branch"
        ).prefetch_related("phases", "project__scopes")

        if user.role == "BRANCH_MANAGER":
            if user.division:
                meeting_jobs = meeting_jobs.filter(project__branch=user.division)
            else:
                meeting_jobs = meeting_jobs.none()
        elif user.role == "PROJECT_MANAGER":
            meeting_jobs = meeting_jobs.filter(project__project_manager=user)

        wb = Workbook()
        ws = wb.active
        ws.title = (str(meeting.meeting_date)[:31] if meeting.meeting_date else "Meeting")

        header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center = Alignment(horizontal="center", vertical="center")

        ws["A1"] = f"Meeting Report - {meeting.meeting_date}"
        ws["A1"].font = Font(bold=True, size=14)

        row = 3
        ws[f"A{row}"] = "Meeting Date:"
        ws[f"B{row}"] = str(meeting.meeting_date)
        row += 1
        ws[f"A{row}"] = "Created By:"
        ws[f"B{row}"] = meeting.created_by.get_full_name() or meeting.created_by.username
        row += 2

        headers = ["Job Number", "Project", "Branch", "PM", "Foreman", "Sat", "Weekends", "Masons", "Labors", "Notes"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        row += 1

        for mj in meeting_jobs:
            p = mj.project
            ws.cell(row=row, column=1, value=p.job_number or "").border = border
            ws.cell(row=row, column=2, value=p.name or "").border = border
            ws.cell(row=row, column=3, value=(p.branch.name if p.branch else "N/A")).border = border
            ws.cell(row=row, column=4, value=(p.project_manager.get_full_name() if p.project_manager else "N/A")).border = border
            ws.cell(row=row, column=5, value=(p.foreman.get_full_name() if p.foreman else "N/A")).border = border
            ws.cell(row=row, column=6, value=("Yes" if mj.saturdays is True else ("No" if mj.saturdays is False else "N/A"))).border = border
            ws.cell(row=row, column=7, value=("Yes" if mj.full_weekends is True else ("No" if mj.full_weekends is False else "N/A"))).border = border
            ws.cell(row=row, column=8, value=mj.masons).border = border
            ws.cell(row=row, column=9, value=mj.labors).border = border
            ws.cell(row=row, column=10, value=mj.notes or "").border = border
            row += 1

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"meeting_{meeting.meeting_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ---------------------------------------------------------
    # Delete restriction
    # ---------------------------------------------------------
    def destroy(self, request, *args, **kwargs):
        user = request.user
        if user.role not in ["ROOT_SUPERADMIN", "SUPERADMIN", "ADMIN"]:
            return Response(
                {"detail": "You do not have permission to delete meetings."},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)
